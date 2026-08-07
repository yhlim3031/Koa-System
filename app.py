from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties

def set_axis_title_horizontal(axis, text):
    axis.title = text
    # Paksa tajuk paksi kekal 0° (mendatar) walaupun label tick di-rotate 90°
    axis.title.tx.rich.bodyPr.rot = 0
    axis.title.tx.rich.bodyPr.vert = "horz"

def rotate_axis_labels(axis, degrees=90):
    # tickLblRot sahaja SENYAP GAGAL (tak serialize ke XML) — mesti guna txPr/bodyPr rot
    # rot dalam unit 60000-per-darjah
    axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=degrees * 60000, vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties()), endParaRPr=CharacterProperties())]
    )

def brighten_marker(series, marker_color="FF0000", line_color="4472C4"):
    # Garisan siri data: kekal terang normal (1.0x) — biru pekat standard
    series.graphicalProperties.line.solidFill = line_color
    series.graphicalProperties.line.width = 19050  # ~1.5pt, garis normal

    # Marker 'x' ialah bentuk TERBUKA (dua garis bersilang) — TIADA kawasan isi.
    # Warna yang sebenarnya nampak ialah stroke/'ln', BUKAN 'solidFill'.
    # Sebab itu solidFill merah tak nampak sebelum ini — ln (outline) kena jadi merah.
    series.marker.symbol = 'circle'
    series.marker.size = 2
    series.marker.graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill=marker_color, w=6350)  # ~2.25pt, tebal & terang
    )
import io

app = Flask(__name__)
CORS(app)

# Warna Excel standard
COLOR_NORMAL_BG = "FFC6EFCE"
COLOR_NORMAL_FONT = "FF006100"
COLOR_WARNING_BG = "FFFFEB9C"
COLOR_WARNING_FONT = "FF9C5700"
COLOR_DANGER_BG = "FFFFC7CE"
COLOR_DANGER_FONT = "FF9C0006"
COLOR_HEADER_BG = "FF4472C4"

def style_cell(cell, is_header=False, status=None):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border
    if is_header:
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF", size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if status:
            if status == "WARNING":
                cell.fill = PatternFill(start_color=COLOR_WARNING_BG, end_color=COLOR_WARNING_BG, fill_type="solid")
                cell.font = Font(bold=True, color=COLOR_WARNING_FONT)
            elif status == "DANGER":
                cell.fill = PatternFill(start_color=COLOR_DANGER_BG, end_color=COLOR_DANGER_BG, fill_type="solid")
                cell.font = Font(bold=True, color=COLOR_DANGER_FONT)
            else:
                cell.fill = PatternFill(start_color=COLOR_NORMAL_BG, end_color=COLOR_NORMAL_BG, fill_type="solid")
                cell.font = Font(bold=True, color=COLOR_NORMAL_FONT)

def style_chart_header(ws, row, text):
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = Font(bold=True, sz=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')

@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    data = request.json
    rows = data.get('data', [])
    date_str = data.get('date_str', 'Tidak diketahui')
    time_str = data.get('time_str', 'Tidak diketahui')

    row_count = len(rows)
    if row_count == 0:
        return "No data", 400

    wb = Workbook()

    ws_temp_data = wb.active
    ws_temp_data.title = "Temperature Data"
    ws_temp_data.append(["LAPORAN DATA SUHU (°C)"])
    ws_temp_data.append([f"Tarikh: {date_str} | Masa: {time_str}"])
    ws_temp_data.merge_cells("A1:B1")
    ws_temp_data.merge_cells("A2:B2")
    ws_temp_data.append(["Masa", "Suhu (°C)", "Status"])
    for cell in ws_temp_data[3]:
        style_cell(cell, is_header=True)
    for row in rows:
        t = row.get('temperature', 0)
        s = row.get('tempStatus', 'NORMAL')
        ws_temp_data.append([row.get('time'), round(t, 1), s])
    for i in range(4, ws_temp_data.max_row + 1):
        status = ws_temp_data.cell(row=i, column=3).value
        for col in range(1, 4):
            style_cell(ws_temp_data.cell(row=i, column=col), status=status)
        ws_temp_data.cell(row=i, column=2).number_format = '0.0'
    ws_temp_data.column_dimensions['A'].width = 30
    ws_temp_data.column_dimensions['B'].width = 15
    ws_temp_data.column_dimensions['C'].width = 20

    ws_temp_chart = wb.create_sheet("Temperature Chart")
    for col in range(1, 6):
        ws_temp_chart.column_dimensions[chr(64 + col)].width = 25
    style_chart_header(ws_temp_chart, 1, "LAPORAN DATA SUHU (°C)")
    style_chart_header(ws_temp_chart, 2, f"Tarikh: {date_str} | Masa: {time_str}")
    ws_temp_chart.merge_cells('A1:E1')
    ws_temp_chart.merge_cells('A2:E2')

    # min_row=3 supaya baris header ("Suhu (°C)") turut disertakan sebagai sumber nama siri
    data_values = Reference(ws_temp_data, min_col=2, min_row=3, max_row=row_count+3)
    data_cats = Reference(ws_temp_data, min_col=1, min_row=4, max_row=row_count+3)

    chart_temp = LineChart()
    chart_temp.title = "Suhu (°C)"
    chart_temp.style = 13
    chart_temp.add_data(data_values, titles_from_data=True)
    chart_temp.set_categories(data_cats)

    brighten_marker(chart_temp.series[0])

    chart_temp.y_axis.title = "Suhu (°C)"
    set_axis_title_horizontal(chart_temp.x_axis, "Masa")
    rotate_axis_labels(chart_temp.x_axis, 90)
    chart_temp.legend = None
    chart_temp.y_axis.scaling.min = 0
    chart_temp.y_axis.scaling.max = 50
    chart_temp.x_axis.auto = False
    chart_temp.x_axis.tickLblPos = 'low'
    chart_temp.width = 27
    chart_temp.height = 13.5
    ws_temp_chart.add_chart(chart_temp, "A4")

    ws_humid_data = wb.create_sheet("Humidity Data")
    ws_humid_data.append(["LAPORAN DATA KELEMBAPAN (%)"])
    ws_humid_data.append([f"Tarikh: {date_str} | Masa: {time_str}"])
    ws_humid_data.merge_cells("A1:B1")
    ws_humid_data.merge_cells("A2:B2")
    ws_humid_data.append(["Masa", "Kelembapan (%)", "Status"])
    for cell in ws_humid_data[3]:
        style_cell(cell, is_header=True)
    for row in rows:
        h = row.get('humidity', 0)
        s = row.get('humidStatus', 'NORMAL')
        ws_humid_data.append([row.get('time'), round(h, 1), s])
    for i in range(4, ws_humid_data.max_row + 1):
        status = ws_humid_data.cell(row=i, column=3).value
        for col in range(1, 4):
            style_cell(ws_humid_data.cell(row=i, column=col), status=status)
        ws_humid_data.cell(row=i, column=2).number_format = '0.0'
    ws_humid_data.column_dimensions['A'].width = 30
    ws_humid_data.column_dimensions['B'].width = 22
    ws_humid_data.column_dimensions['C'].width = 20

    ws_humid_chart = wb.create_sheet("Humidity Chart")
    for col in range(1, 6):
        ws_humid_chart.column_dimensions[chr(64 + col)].width = 25
    style_chart_header(ws_humid_chart, 1, "LAPORAN DATA KELEMBAPAN (%)")
    style_chart_header(ws_humid_chart, 2, f"Tarikh: {date_str} | Masa: {time_str}")
    ws_humid_chart.merge_cells('A1:E1')
    ws_humid_chart.merge_cells('A2:E2')

    data_humid_values = Reference(ws_humid_data, min_col=2, min_row=3, max_row=row_count+3)

    chart_humid = LineChart()
    chart_humid.title = "Kelembapan (%)"
    chart_humid.style = 12
    chart_humid.add_data(data_humid_values, titles_from_data=True)
    chart_humid.set_categories(data_cats)

    brighten_marker(chart_humid.series[0])

    chart_humid.y_axis.title = "Kelembapan (%)"
    set_axis_title_horizontal(chart_humid.x_axis, "Masa")
    rotate_axis_labels(chart_humid.x_axis, 90)
    chart_humid.legend = None
    chart_humid.y_axis.scaling.min = 0
    chart_humid.y_axis.scaling.max = 100
    chart_humid.x_axis.auto = False
    chart_humid.x_axis.tickLblPos = 'low'
    chart_humid.width = 27
    chart_humid.height = 13.5
    ws_humid_chart.add_chart(chart_humid, "A4")

    ws_gas_data = wb.create_sheet("Gas Data")
    ws_gas_data.append(["LAPORAN DATA GAS (ADC)"])
    ws_gas_data.append([f"Tarikh: {date_str} | Masa: {time_str}"])
    ws_gas_data.merge_cells("A1:B1")
    ws_gas_data.merge_cells("A2:B2")
    ws_gas_data.append(["Masa", "Gas (ADC)", "Status"])
    for cell in ws_gas_data[3]:
        style_cell(cell, is_header=True)
    for row in rows:
        g = row.get('gas', 0)
        s = row.get('gasStatus', 'NORMAL')
        ws_gas_data.append([row.get('time'), round(g), s])
    for i in range(4, ws_gas_data.max_row + 1):
        status = ws_gas_data.cell(row=i, column=3).value
        for col in range(1, 4):
            style_cell(ws_gas_data.cell(row=i, column=col), status=status)
    ws_gas_data.column_dimensions['A'].width = 30
    ws_gas_data.column_dimensions['B'].width = 14
    ws_gas_data.column_dimensions['C'].width = 20

    ws_gas_chart = wb.create_sheet("Gas Chart")
    for col in range(1, 6):
        ws_gas_chart.column_dimensions[chr(64 + col)].width = 25
    style_chart_header(ws_gas_chart, 1, "LAPORAN DATA GAS (ADC)")
    style_chart_header(ws_gas_chart, 2, f"Tarikh: {date_str} | Masa: {time_str}")
    ws_gas_chart.merge_cells('A1:E1')
    ws_gas_chart.merge_cells('A2:E2')

    data_gas_values = Reference(ws_gas_data, min_col=2, min_row=3, max_row=row_count+3)

    chart_gas = LineChart()
    chart_gas.title = "Gas (ADC)"
    chart_gas.style = 11
    chart_gas.add_data(data_gas_values, titles_from_data=True)
    chart_gas.set_categories(data_cats)

    brighten_marker(chart_gas.series[0])

    chart_gas.y_axis.title = "Gas (ADC)"
    set_axis_title_horizontal(chart_gas.x_axis, "Masa")
    rotate_axis_labels(chart_gas.x_axis, 90)
    chart_gas.legend = None
    chart_gas.y_axis.scaling.min = 0
    chart_gas.y_axis.scaling.max = 1000
    chart_gas.x_axis.auto = False
    chart_gas.x_axis.tickLblPos = 'low'
    chart_gas.width = 27
    chart_gas.height = 13.5
    ws_gas_chart.add_chart(chart_gas, "A4")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name="LAPORAN SISTEM MONITORING PENGGERA KEBAKARAN.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)